import datetime
import os

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from condoparser.jpa import db
from condoparser.jpa.Queries import GET_OFFERS_THAT_NO_LONGER_EXISTS_QUERY, GET_OFFERS_THAT_GOT_UPDATED, GET_NEW_OFFERS


os.makedirs(
    os.getenv("PARSER_REPORT_LOCATION") +
    "/" +
    datetime.datetime.now().strftime('%Y-%m-%d'),
    exist_ok=True)


def processNewOffers(newOffers):
    ws = wb.active
    ws.title = "newOffers"
    colnames = [
        'Zone',
        'Condo',
        'Offer',
        'NoOfRooms',
        'NoOfBathrooms',
        'ConstructionDate',
        'BalconyArea',
        'UsableArea',
        'CurrentPrice']
    data = []
    for row in newOffers:
        data.append(
            (row.condo_zone,
             row.condo_name,
             row.url,
             row.room_no,
             row.bathroom_no,
             row.construction_date,
             row.terase_usable_area,
             row.usable_area,
             row.current_price))

    constructSheet(ws, colnames, data, str(newOffers.rowcount + 1))


def processDeletedOffers(deletedOffers):
    ws = wb.create_sheet("Delete Offers")
    ws.title = "deletedOffers"
    colnames = [
        'Zone',
        'Condo',
        'Offer',
        'NoOfRooms',
        'NoOfBathrooms',
        'ConstructionDate',
        'BalconyArea',
        'UsableArea',
        'CurrentPrice']
    data = []
    for row in deletedOffers:
        data.append(
            (row.condo_zone,
             row.condo_name,
             row.url,
             row.room_no,
             row.bathroom_no,
             row.construction_date,
             row.terase_usable_area,
             row.usable_area,
             row.current_price))
    constructSheet(ws, colnames, data, str(deletedOffers.rowcount + 1))


def processUpdatedOffers(updatedOffers):
    ws = wb.create_sheet("Delete Offers")
    ws.title = "updatedOffers"
    colnames = [
        'Zone',
        'Condo',
        'Offer',
        'OLD NoOfRooms',
        'NEW NoOfRooms',
        'OLD NoOfBathrooms',
        'NEW NoOfBathrooms',
        'OLD ConstructionDate',
        'NEW ConstructionDate',
        'OLD BalconyArea',
        'NEW BalconyArea',
        'OLD UsableArea',
        'NEW UsableArea',
        'OLD CurrentPrice',
        'NEW CurrentPrice']
    rowCount = 1
    data = []
    for row in updatedOffers:
        rowCount += 1
        data.append(
            styled_cells(
                ws,
                (row.condo_zone,
                 row.condo_name,
                 row.url,
                 row.old_room_no,
                 row.new_room_no,
                 row.old_bathroom_no,
                 row.new_bathroom_no,
                 row.old_construction_date,
                 row.new_construction_date,
                 row.old_terase_usable_area,
                 row.new_terase_usable_area,
                 row.old_usable_area,
                 row.new_usable_area,
                 row.old_current_price,
                 row.new_current_price),
                rowCount))
    constructSheet(ws, colnames, data, str(updatedOffers.rowcount + 1))


def styled_cells(ws, data, row):
    for idx, val in enumerate(data):
        if idx < 3:
            yield Cell(ws, column=chr(idx + 65), row=row, value=val)
        elif idx % 2 == 1:
            defineColorForCell
            c = defineColorForCell(
                ws, data[idx], data[idx + 1], data[idx], row, idx, 'e5383b')
            yield c
        elif idx % 2 == 0:
            defineColorForCell
            c = defineColorForCell(
                ws, data[idx - 1], data[idx], data[idx], row, idx, '52b788')
            yield c


def defineColorForCell(ws, oldValue, newValue, val, row, column, color):
    cell = Cell(ws, column=chr(column + 65), row=row, value=val)
    if (oldValue != newValue):
        cell.fill = PatternFill(start_color=color, fill_type="solid")
    return cell


def resizeSheets(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except BaseException:
                pass
        adjusted_width = (max_length + 2) * 1
        worksheet.column_dimensions[column].width = adjusted_width


def constructSheet(ws, headerColumns, data, count):
    ws.append(headerColumns)

    for row in data:
        ws.append(row)

    tab = Table(displayName=ws.title, ref="A1:" +
                chr(len(headerColumns) + 64) + count)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    resizeSheets(ws)


if __name__ == '__main__':
    # get data from DB
    deletedOffers = db.session.execute(GET_OFFERS_THAT_NO_LONGER_EXISTS_QUERY)
    modifiedOffers = db.engine.execute(GET_OFFERS_THAT_GOT_UPDATED)
    newOffers = db.engine.execute(GET_NEW_OFFERS)

    wb = Workbook()

    # process data / save results
    processNewOffers(newOffers)
    processDeletedOffers(deletedOffers)
    processUpdatedOffers(modifiedOffers)

    wb.save(os.getenv("PARSER_REPORT_LOCATION") +
            "/" +
            datetime.datetime.now().strftime('%Y-%m-%d') +
            "/condo-parse-results" +
            str(datetime.datetime.now()).replace(' ', '^') +
            ".xlsx")
